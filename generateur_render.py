"""
Generateur COMPLET de dossier de candidature pour Render.
Utilise l'API Claude directement (httpx) sans les dependances lourdes
(pdfplumber, pymupdf, playwright, python-docx, etc.).

Genere un dossier complet :
1. Analyse Go/No-Go
2. Memoire technique
3. Lettre de candidature
4. BPU / DPGF
5. Planning previsionnel
6. CV Formateurs
7. DC1 / DC2
8. Checklist de soumission
"""

import os
import re
import json
import logging
from pathlib import Path
from datetime import datetime

logger = logging.getLogger("ao_hunter.generateur_render")

DASHBOARD_DIR = Path(__file__).parent
DOSSIERS_DIR = DASHBOARD_DIR / "dossiers_generes"
DOSSIERS_INDEX = DASHBOARD_DIR / "dossiers_index.json"

# Cle API depuis variable d'environnement
API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# Modeles
MODELE_PRINCIPAL = "claude-sonnet-4-20250514"
MODELE_LEGER = "claude-sonnet-4-20250514"  # Haiku non dispo sur cette cle, fallback Sonnet

# Infos entreprise integrees (extrait du config.yaml)
ENTREPRISE = {
    "nom": "Almera",
    "raison_sociale": "AI MENTOR",
    "siret": "98900455100010",
    "nda": "11757431975",
    "forme_juridique": "SASU",
    "adresse": "25 rue Campagne Premiere, 75014 Paris",
    "representant": "Mickael Bertolla, President",
    "site_web": "almera.one",
    "email": "contact@almera.one",
    "telephone": "+33686680611",
    "certifications": [
        "Qualiopi (actions de formation)",
        "RS6776 France Competences (IA generative)",
        "Activateur France Num",
        "Membre Hub France IA",
        "Membre French Tech",
    ],
    "chiffres": "2000+ personnes formees, 50+ entreprises, note 4.9/5 Google",
    "formateur_principal": "Mickael Bertolla - Ingenieur Mines Saint-Etienne, MSc Skema (USA), auteur 'L'IA et la generation de texte' (Ed. ENI, best-seller 2025)",
    "formateurs": [
        {"nom": "Mickael Bertolla", "role": "President & Formateur principal", "specialites": "IA generative, ChatGPT, Claude, Agents IA, Strategie IA, Prompt engineering", "formation": "Ingenieur Mines Saint-Etienne, MSc Skema Business School (USA)", "experience": "4 ans formation IA, 2000+ personnes formees, auteur Ed. ENI"},
        {"nom": "Charles Lerminiaux", "role": "Consultant Data et IA senior", "specialites": "Data science, gouvernance donnees, cadrage cas d'usage IA pour COMEX, feuille de route IA", "formation": "Ingenieur ISAE Supaero, MS Management Innovation TBS, CDMP Associate", "experience": "17 ans, ex-Directeur conseil Aqsone, Product Manager IA Airbus, Sopra Steria"},
        {"nom": "Guillaume Lanz", "role": "Consultant IA / CEO Transition IA", "specialites": "IA generative, Microsoft Copilot (20+ ateliers), prompting avance, LinkedIn IA", "formation": "MBA Manager Produits et Marketing ESG Paris, certifie Google GenAI", "experience": "3 ans, 1500+ professionnels formes, conferencier BPI France / AI Summit"},
        {"nom": "Romy Ozier-Lafontaine", "role": "Formatrice IA et design numerique", "specialites": "IA generative, SEO, accessibilite RGAA, WordPress/Webflow, UX Design", "formation": "Dev Web 3W Academy, jury habilite PIX", "experience": "18 ans, fondatrice Digital Ladies, 250+ personnes formees"},
        {"nom": "Yann Cabon", "role": "Formateur Expert IA", "specialites": "IA generative (texte/image/video), creation contenu IA, integration IA processus RH", "formation": "Master RH Ecofac Business School", "experience": "2 ans formation IA, ex-SNCF Voyageurs, 900+ contenus IA generes"},
        {"nom": "Stephanie Rodrigues", "role": "Consultante et Formatrice IA", "specialites": "IA generative, automatisation, transformation digitale", "formation": "Consultante certifiee", "experience": "Formatrice IA en entreprise"},
    ],
    "grille_tarifaire": {
        "journee_presentiel": "1 500 EUR HT",
        "journee_distanciel": "1 200 EUR HT",
        "journee_preparation": "800 EUR HT",
        "heure_suivi": "150 EUR HT",
        "coaching_individuel": "1 800 EUR HT / jour",
        "elearning_par_stagiaire": "200 EUR HT",
        "certification_rs6776": "500 EUR HT / stagiaire",
        "diagnostic_audit": "1 500 EUR HT / jour",
    },
}


def _appel_claude(prompt: str, max_tokens: int = 4000, modele: str = None) -> str:
    """Appelle l'API Claude et retourne le texte de la reponse."""
    import httpx

    model = modele or MODELE_PRINCIPAL

    with httpx.Client(timeout=300) as client:
        resp = client.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": model,
                "max_tokens": max_tokens,
                "messages": [{"role": "user", "content": prompt}],
            },
        )
        resp.raise_for_status()
        data = resp.json()

    return data["content"][0]["text"]


def _formater_date(raw: str) -> str:
    """Formate une date ISO en date lisible."""
    if not raw or raw == "None":
        return "Non precisee"
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        mois = ["janvier", "fevrier", "mars", "avril", "mai", "juin",
                "juillet", "aout", "septembre", "octobre", "novembre", "decembre"]
        return f"{dt.day} {mois[dt.month - 1]} {dt.year} a {dt.hour}h{dt.minute:02d}"
    except Exception:
        return str(raw).split("T")[0] if "T" in str(raw) else str(raw)


def _bloc_infos_ao(ao: dict) -> str:
    """Bloc d'infos AO reutilise dans tous les prompts."""
    budget = ao.get('budget_estime')
    budget_txt = f"{budget:,.0f} EUR".replace(",", " ") if budget else "Non communique"
    date_lim = _formater_date(str(ao.get('date_limite', '')))

    return f"""Titre : {ao.get('titre', 'Non precise')}
Acheteur : {ao.get('acheteur', 'Non precise')}
Description : {ao.get('description', 'Non disponible')}
Reference/ID : {ao.get('id', '')}
Budget estime : {budget_txt}
Type : {ao.get('type_marche', 'Non precise')}
Procedure : {ao.get('type_procedure', 'Non precise')}
Lieu : {ao.get('region', ao.get('lieu_execution', 'Non precise'))}
Duree : {ao.get('duree_mois', 'Non precisee')} mois
Date limite : {date_lim}
Criteres : {ao.get('criteres_attribution', 'Non precises')}"""


def _bloc_infos_entreprise() -> str:
    """Bloc d'infos entreprise reutilise dans tous les prompts."""
    ent = ENTREPRISE
    return f"""Raison sociale : {ent['raison_sociale']}
Nom commercial : {ent['nom']}
Forme juridique : {ent['forme_juridique']}
SIRET : {ent['siret']}
NDA : {ent['nda']}
Adresse : {ent['adresse']}
Representant : {ent['representant']}
Telephone : {ent['telephone']}
Email : {ent['email']}
Site web : {ent['site_web']}
Certifications : {', '.join(ent['certifications'])}
Chiffres cles : {ent['chiffres']}
Formateur principal : {ent['formateur_principal']}"""


# --- Helpers ---

def _extraire_criteres_attribution(ao: dict, dce_texte: str = "") -> list[dict]:
    """Extrait les criteres d'attribution de l'AO et du DCE.

    Retourne une liste triee par poids decroissant :
    [{"nom": "Valeur technique", "poids_pct": 60, "sous_criteres": [
        {"nom": "Methodologie", "poids_pct": 30}, ...
    ]}, ...]
    """
    criteres = []
    texte_criteres = ""

    # Source 1 : champ criteres_attribution de l'AO
    crit_ao = ao.get("criteres_attribution", "")
    if isinstance(crit_ao, str) and crit_ao and crit_ao != "Non precises":
        texte_criteres += " " + crit_ao
    elif isinstance(crit_ao, list):
        for c in crit_ao:
            if isinstance(c, dict):
                criteres.append({
                    "nom": c.get("nom", c.get("name", "")),
                    "poids_pct": c.get("poids_pct", c.get("poids", c.get("weight", 0))),
                    "sous_criteres": c.get("sous_criteres", []),
                })
            elif isinstance(c, str):
                texte_criteres += " " + c

    # Source 2 : texte du DCE
    if dce_texte:
        # Chercher la section criteres dans le DCE
        for pattern in [
            r"crit[eè]res?\s+d['']attribution",
            r"crit[eè]res?\s+de\s+jugement",
            r"crit[eè]res?\s+d['']analyse",
            r"pond[eé]ration",
            r"notation\s+des\s+offres",
        ]:
            match = re.search(pattern, dce_texte, re.IGNORECASE)
            if match:
                # Extraire un bloc de 2000 chars autour du match
                start = max(0, match.start() - 200)
                end = min(len(dce_texte), match.end() + 2000)
                texte_criteres += " " + dce_texte[start:end]
                break

    # Si on a deja des criteres structures, les retourner
    if criteres:
        criteres.sort(key=lambda c: c["poids_pct"], reverse=True)
        return criteres

    # Parser le texte pour extraire criteres et ponderations
    if not texte_criteres.strip():
        return []

    # Pattern : "Nom du critere XX%" ou "Nom du critere : XX %" ou "Nom (XX%)"
    # Ex: "Valeur technique 60%" ou "Prix : 40 %"
    pattern_critere = re.compile(
        r'([A-ZÀ-Ÿa-zà-ÿ][A-ZÀ-Ÿa-zà-ÿ\s/\-\']+?)\s*[:(\s]\s*(\d{1,3})\s*%',
        re.UNICODE
    )
    matches = pattern_critere.findall(texte_criteres)

    for nom_raw, poids_raw in matches:
        nom = nom_raw.strip().rstrip(" :(")
        poids = int(poids_raw)
        # Filtrer les faux positifs (poids hors limites, noms trop courts)
        if poids < 1 or poids > 100 or len(nom) < 3:
            continue
        # Eviter les doublons
        if any(c["nom"].lower() == nom.lower() for c in criteres):
            continue
        criteres.append({"nom": nom, "poids_pct": poids, "sous_criteres": []})

    # Detecter les sous-criteres : "dont Xxx YY%, Yyy ZZ%"
    pattern_dont = re.compile(
        r'dont\s+(.*?)(?:\.|;|\n|$)',
        re.IGNORECASE | re.UNICODE
    )
    for match in pattern_dont.finditer(texte_criteres):
        bloc_sous = match.group(1)
        sous_matches = re.findall(
            r'([A-ZÀ-Ÿa-zà-ÿ][A-ZÀ-Ÿa-zà-ÿ\s/\-\']+?)\s*(\d{1,3})\s*%',
            bloc_sous
        )
        if sous_matches:
            # Trouver le critere parent (le plus proche avant "dont")
            pos_dont = match.start()
            parent = None
            for c in criteres:
                if c["nom"].lower() in texte_criteres[:pos_dont + 50].lower():
                    parent = c
            # Si pas de parent trouve, attribuer au premier critere non-prix
            if not parent:
                for c in criteres:
                    if "prix" not in c["nom"].lower():
                        parent = c
                        break
            if parent:
                for sc_nom, sc_poids in sous_matches:
                    sc_nom = sc_nom.strip()
                    if len(sc_nom) >= 3 and int(sc_poids) <= parent["poids_pct"]:
                        parent["sous_criteres"].append({
                            "nom": sc_nom,
                            "poids_pct": int(sc_poids),
                        })

    criteres.sort(key=lambda c: c["poids_pct"], reverse=True)
    return criteres


# --- Generation de chaque piece du dossier ---

def _objectif_mots(criteres: list) -> int:
    """Calcule l'objectif de mots du memoire selon le poids de la valeur technique."""
    if not criteres:
        return 3000
    # Chercher le poids du critere technique (tout sauf prix)
    poids_technique = 0
    for c in criteres:
        nom = c.get("nom", "").lower()
        if "prix" not in nom and "cout" not in nom and "tarif" not in nom:
            poids_technique += c.get("poids_pct", 0)
    # Plus le poids technique est eleve, plus le memoire doit etre long
    if poids_technique >= 60:
        return 5000
    elif poids_technique >= 40:
        return 4000
    elif poids_technique >= 20:
        return 3000
    else:
        return 2500  # Peu de poids technique = memoire plus court


def _seuil_dynamique(ao: dict) -> float:
    """Calcule un seuil de pertinence dynamique selon le type d'acheteur et le budget."""
    acheteur = (ao.get('acheteur', '') or '').lower()
    budget = ao.get('budget_estime', 0) or 0

    # Type d'acheteur
    mots_public = ['ministere', 'region', 'departement', 'commune', 'mairie',
                   'collectivite', 'cnrs', 'universite', 'inserm', 'chu', 'hopital',
                   'crous', 'rectorat', 'prefecture', 'agence', 'etablissement public',
                   'syndicat mixte', 'communaute', 'metropole', 'conseil']
    mots_education = ['universite', 'ecole', 'lycee', 'college', 'crous', 'rectorat',
                      'education', 'enseignement', 'campus', 'iut', 'bts']

    is_public = any(m in acheteur for m in mots_public)
    is_education = any(m in acheteur for m in mots_education)

    # Seuil de base
    if is_education:
        seuil = 0.55  # Missions alignees, souvent moins concurrentielles
    elif is_public:
        seuil = 0.60  # Bon historique de gain sur le public
    else:
        seuil = 0.70  # Prive = plus exigeant sur le fit

    # Ajustement budget
    if 5000 <= budget <= 70000:
        seuil -= 0.05  # Zone ideale Almera = plus permissif
    elif budget > 100000:
        seuil += 0.05  # Gros marche = plus de concurrence

    return round(max(0.40, min(0.85, seuil)), 2)


def _generer_memoire(ao: dict, type_presta: str, dce_texte: str = "", personnalisation: dict = None, decp: dict = None) -> str:
    """Genere le memoire technique (piece maitresse)."""
    dce_bloc = ""
    if dce_texte:
        dce_bloc = f"""

=== EXTRAIT DU DCE ===
Utilise ces extraits du Dossier de Consultation des Entreprises pour personnaliser
au maximum le memoire (reprendre les termes exacts, repondre aux exigences detectees) :
{dce_texte[:8000]}
==="""

    # Feature 1 : Structure en miroir des criteres d'attribution
    criteres = _extraire_criteres_attribution(ao, dce_texte)
    criteres_bloc = ""
    structure_bloc = ""
    if criteres:
        criteres_lignes = []
        for c in criteres:
            ligne = f"- {c['nom']} ({c['poids_pct']}%)"
            if c.get("sous_criteres"):
                for sc in c["sous_criteres"]:
                    ligne += f"\n  - {sc['nom']} ({sc['poids_pct']}%)"
            criteres_lignes.append(ligne)

        criteres_bloc = f"""

=== CRITERES D'ATTRIBUTION DETECTES ===
{chr(10).join(criteres_lignes)}
==="""

        # Generer la structure imposee du memoire
        sections = []
        for i, c in enumerate(criteres, 1):
            if "prix" in c["nom"].lower():
                continue  # Le prix est traite dans le BPU, pas dans le memoire
            if c.get("sous_criteres"):
                sections.append(f"## {i}. {c['nom']} ({c['poids_pct']}%)")
                for j, sc in enumerate(c["sous_criteres"], 1):
                    sections.append(f"### {i}.{j}. {sc['nom']} ({sc['poids_pct']}%)")
            else:
                sections.append(f"## {i}. {c['nom']} ({c['poids_pct']}%)")

        structure_bloc = f"""

=== STRUCTURE OBLIGATOIRE DU MEMOIRE ===
IMPORTANT : Le memoire DOIT reprendre EXACTEMENT ces sections dans cet ordre,
en utilisant les INTITULES EXACTS des criteres d'attribution.
Le volume de chaque section doit etre proportionnel a son poids dans la notation.

{chr(10).join(sections)}
==="""

    # Feature 2 : Personnalisation acheteur
    perso_bloc = ""
    if personnalisation:
        from personnalisation_acheteur import bloc_personnalisation_prompt
        perso_bloc = bloc_personnalisation_prompt(personnalisation)

    # Feature : Memoire adaptative - injection des modeles gagnants similaires
    adaptatif_bloc = ""
    try:
        from memoire_adaptative import trouver_modeles_similaires, generer_prompt_adaptatif
        modeles = trouver_modeles_similaires(ao, n=3)
        if modeles:
            adaptatif_bloc = generer_prompt_adaptatif(ao, modeles)
            logger.info(f"Memoire adaptative: {len(modeles)} modeles injectes (meilleur score: {modeles[0]['similarite']})")
    except Exception as e:
        logger.warning(f"Memoire adaptative: erreur recherche modeles: {e}")

    # Feature : Lecons des AO gagnes (intelligence adaptative)
    lecons_bloc = ""
    try:
        from memoire_adaptative import appliquer_lecons
        lecons_bloc = appliquer_lecons(ao)
        if lecons_bloc:
            logger.info("Memoire adaptative: lecons d'AO gagnes injectees dans le prompt")
    except Exception as e:
        logger.warning(f"Memoire adaptative: erreur lecons: {e}")

    # Feature : Intelligence concurrentielle (analyse concurrents + contre-arguments)
    concurrence_bloc = ""
    try:
        from intelligence_concurrentielle import analyser_concurrents_ao, concurrents_par_defaut, enrichir_prompt_memoire
        noms = concurrents_par_defaut()
        analyse_conc = analyser_concurrents_ao(noms, ao=ao)
        prompt_conc = analyse_conc.get("prompt_memoire", "")
        if prompt_conc:
            concurrence_bloc = f"""

=== {prompt_conc} ==="""
            logger.info(f"Intelligence concurrentielle: {len(analyse_conc.get('concurrents', []))} concurrents analyses")
    except Exception as e:
        logger.warning(f"Intelligence concurrentielle non disponible: {e}")

    # Contexte DECP (données marché réelles)
    decp_bloc = ""
    if decp and decp.get("nb_marches_trouves", 0) > 0:
        decp_bloc = f"""

=== DONNEES MARCHE REELLES (source: DECP data.gouv.fr, {decp['nb_marches_trouves']} marches similaires) ===
- Prix median des marches similaires: {decp['budget']['median']} EUR
- Fourchette de prix: {decp['budget']['fourchette_recommandee'][0]} - {decp['budget']['fourchette_recommandee'][1]} EUR
- Nombre median de candidats: {decp['concurrence']['nb_offres_median']}
- Concurrents frequents: {', '.join(t['nom'] for t in decp.get('titulaires_frequents', [])[:5])}
Utilisez ces donnees pour calibrer le positionnement tarifaire et les arguments differenciants.
==="""

    prompt = f"""Tu es un expert en reponse aux appels d'offres publics francais.
Genere un MEMOIRE TECHNIQUE complet et professionnel pour cet appel d'offres.

## APPEL D'OFFRES
{_bloc_infos_ao(ao)}

## TYPE DE PRESTATION DETECTE : {type_presta}

## ENTREPRISE CANDIDATE
{_bloc_infos_entreprise()}
{dce_bloc}{criteres_bloc}{structure_bloc}{perso_bloc}{adaptatif_bloc}{lecons_bloc}{concurrence_bloc}{decp_bloc}

## INSTRUCTIONS
1. Redige un memoire technique COMPLET en francais (minimum {_objectif_mots(criteres)} mots)
2. {"Structure le memoire en MIROIR EXACT des criteres d'attribution ci-dessus, en respectant l'ordre et les intitules." if criteres else f'Structure avec sections standard pour "{type_presta}"'}
3. Personnalise chaque section au contexte de l'acheteur
4. Utilise UNIQUEMENT les vraies references et certifications d'Almera
5. N'invente JAMAIS de noms de formateurs ou de references clients
6. Propose des indicateurs de performance (KPIs) concrets et mesurables
7. Inclus un planning previsionnel realiste avec jalons
8. Si Formation : objectifs pedagogiques SMART, methodes actives detaillees (ateliers, cas pratiques, mises en situation), supports numeriques innovants, modalites d'evaluation formative et sommative
9. Si Consulting/AMO : methodologie en 4 etapes (diagnostic, feuille de route, accompagnement, deploiement) avec livrables concrets par etape
10. Sois precis, concret, quantifie. Pas de phrases generiques. Chaque argument doit etre illustre par un exemple reel.
11. Format Markdown avec titres # ## ###
12. Pour chaque section, inclus au moins une reference client pertinente avec resultats chiffres
{"13. Le VOLUME de chaque section doit etre PROPORTIONNEL au poids du critere (ex: critere a 30% = ~30% du memoire)" if criteres else ""}
14. IMPORTANT : Le memoire DOIT se terminer proprement avec une conclusion. Ne depasse pas {_objectif_mots(criteres) + 1500} mots. Termine TOUJOURS par une section "CONCLUSION" de 3-5 phrases resumant les points forts de la candidature.
15. DIFFERENCIATION CONCURRENTIELLE : Pour chaque point cle, explique en quoi Almera se distingue des autres organismes de formation (approche en 4 etapes, certification RS6776 rare, personnalisation 100%, reseau de 10 formateurs specialises, auteur publie chez ENI)
16. PREUVES CONCRETES : Chaque affirmation doit etre etayee par un chiffre reel (2000+ personnes formees, 50+ entreprises, 4.9/5 Google, 80+ personnes chez Havas, etc.)
17. MAPPING RC : Si des criteres d'attribution sont detectes, commence chaque section par "En reponse au critere [nom du critere] :" pour faciliter la notation par l'acheteur
18. ENGAGEMENT QUALITE : Inclus systematiquement les engagements : taux de satisfaction > 4.5/5, taux de completion > 95%, suivi post-formation 3 mois, hotline 48h"""

    # Adapter max_tokens a l'objectif de mots (1 mot ~ 1.5 tokens en francais)
    objectif = _objectif_mots(criteres)
    max_tok = max(16000, int(objectif * 3.0))
    return _appel_claude(prompt, max_tokens=max_tok)


def _detecter_persona_acheteur(acheteur: str) -> str:
    """Detecte le type d'acheteur et retourne les instructions de ton pour le prompt."""
    acheteur_lower = (acheteur or '').lower()

    mots_institutionnel = ['ministere', 'region', 'departement', 'commune', 'mairie',
                           'collectivite', 'cnrs', 'universite', 'inserm', 'chu', 'hopital']
    mots_prive = ['sas', 'sa ', 'sarl', 'group', 'corp']

    if any(m in acheteur_lower for m in mots_institutionnel):
        return ("Tone detecte : FORMEL/INSTITUTIONNEL\n"
                "- Utiliser le vouvoiement strict et un registre soutenu\n"
                "- References aux cadres reglementaires (CCP, Qualiopi, France Competences)\n"
                "- Insister sur la conformite, les garanties et les engagements de service public\n"
                "- Vocabulaire : 'nous avons l'honneur', 'conformement a', 'nous nous engageons'")
    elif any(m in acheteur_lower for m in mots_prive):
        return ("Tone detecte : PROFESSIONNEL/DYNAMIQUE\n"
                "- Ton professionnel mais dynamique, axe resultats et ROI\n"
                "- Mettre en avant l'agilite, la personnalisation et les retours concrets\n"
                "- Vocabulaire : 'partenaire', 'performance', 'valeur ajoutee', 'impact mesurable'")
    else:
        return ("Tone detecte : PROFESSIONNEL\n"
                "- Ton professionnel equilibre\n"
                "- Combiner rigueur institutionnelle et orientation resultats\n"
                "- Vocabulaire adapte au contexte de l'acheteur")


def _generer_lettre(ao: dict, personnalisation: dict = None) -> str:
    """Genere la lettre de candidature."""
    ent = ENTREPRISE

    # Feature 2 : Personnalisation acheteur
    perso_bloc = ""
    if personnalisation:
        from personnalisation_acheteur import bloc_personnalisation_prompt
        perso_bloc = bloc_personnalisation_prompt(personnalisation)

    prompt = f"""Redige une LETTRE DE CANDIDATURE professionnelle et COMPLETE pour cet appel d'offres.

=== APPEL D'OFFRES ===
{_bloc_infos_ao(ao)}

=== INFORMATIONS ENTREPRISE CANDIDATE ===
{_bloc_infos_entreprise()}
{perso_bloc}

=== REGLES IMPERATIVES ===
- NE LAISSE AUCUN TEXTE ENTRE CROCHETS []. Toutes les informations sont ci-dessus.
- NE LAISSE AUCUN CHAMP VIDE ou "A completer". Utilise les infos fournies.
- La lettre doit etre 100% prete a imprimer et signer, sans aucune modification necessaire.
- Utilise la date du jour : {datetime.now().strftime('%d/%m/%Y')}

=== PERSONA ACHETEUR ===
{_detecter_persona_acheteur(ao.get('acheteur', ''))}

=== CONTENU ===
- Adresser a l'acheteur ({ao.get('acheteur', '')})
- Mentionner l'objet exact du marche
- Presenter l'entreprise et sa qualification
- Declarer sur l'honneur l'absence de cas d'exclusion (art. L.2141-1 a L.2141-5 CCP)
- Attester etre en regle vis-a-vis obligations fiscales et sociales
- Mentionner les certifications (Qualiopi, RS6776)
- Faire reference a 1-2 exigences specifiques extraites de la description de l'AO pour montrer la lecture attentive du cahier des charges
- Inclure l'engagement : "Mickael Bertolla assurera personnellement la direction pedagogique"
- Inclure l'engagement : "Almera s'engage sur un taux de satisfaction superieur a 4.5/5"
- Terminer par un paragraphe de valeur ajoutee avec 3 resultats quantifies (ex: 2000+ personnes formees, 50+ entreprises accompagnees, note 4.9/5 Google)
- Signature par le representant legal avec nom, titre, date
{"- Adapter le ton et le vocabulaire au type d'acheteur (" + personnalisation['type_acheteur'] + ")" if personnalisation else ""}

Format : Lettre professionnelle formelle, prete a imprimer. Adapter le ton selon la PERSONA ACHETEUR detectee ci-dessus."""

    return _appel_claude(prompt, max_tokens=2000)


def _generer_bpu(ao: dict, dce_texte: str = "", decp: dict = None) -> str:
    """Genere le BPU/DPGF avec recommandation de prix dynamique."""
    ent = ENTREPRISE
    grille = "\n".join(f"- {k.replace('_', ' ').title()}: {v}" for k, v in ent["grille_tarifaire"].items())

    # Recommandation de prix dynamique
    reco_bloc = ""
    strategie_commentaire = ""
    try:
        from modeles_prix import recommander_prix
        reco = recommander_prix(ao)
        reco_bloc = f"""

=== RECOMMANDATION DE PRIX (basee sur l'historique) ===
TJM recommande : {reco['tjm_recommande']} EUR/jour
Fourchette competitive : {reco['fourchette'][0]} - {reco['fourchette'][1]} EUR/jour
Prix total estime : {reco['prix_total_estime']} EUR HT
Nb jours estime : {reco['nb_jours_estime']}
Strategie : {reco['strategie']}
Type prestation : {reco['type_prestation']}
IMPORTANT : Utilise le TJM recommande comme base pour les prix unitaires.
==="""
        strategie_commentaire = (
            f"\n\n> **Strategie tarifaire retenue : {reco['strategie'].upper()}** "
            f"| TJM cible : {reco['tjm_recommande']} EUR/jour "
            f"| Fourchette : {reco['fourchette'][0]}-{reco['fourchette'][1]} EUR/jour"
        )
        logger.info(f"BPU: recommandation prix TJM={reco['tjm_recommande']}, strategie={reco['strategie']}")
    except Exception as e:
        logger.warning(f"Recommandation prix indisponible: {e}")

    dce_bloc = ""
    if dce_texte:
        dce_bloc = f"""

=== EXTRAIT DU DCE ===
Utilise ces extraits pour adapter les postes de prix, quantites et unites
aux exigences exactes du cahier des charges :
{dce_texte[:8000]}
==="""

    # Contexte DECP (données marché réelles)
    decp_bloc = ""
    if decp and decp.get("nb_marches_trouves", 0) > 0:
        decp_bloc = f"""

=== DONNEES MARCHE REELLES (source: DECP data.gouv.fr, {decp['nb_marches_trouves']} marches similaires) ===
- Prix median des marches similaires: {decp['budget']['median']} EUR
- Fourchette de prix: {decp['budget']['fourchette_recommandee'][0]} - {decp['budget']['fourchette_recommandee'][1]} EUR
- Nombre median de candidats: {decp['concurrence']['nb_offres_median']}
- Concurrents frequents: {', '.join(t['nom'] for t in decp.get('titulaires_frequents', [])[:5])}
IMPORTANT : Calibrez les prix en tenant compte de ces donnees marche reelles.
==="""

    prompt = f"""Genere un BORDEREAU DE PRIX UNITAIRES (BPU) et une DECOMPOSITION DU PRIX GLOBAL FORFAITAIRE (DPGF).

=== APPEL D'OFFRES ===
{_bloc_infos_ao(ao)}

=== GRILLE TARIFAIRE ALMERA ===
{grille}

=== INFORMATIONS ENTREPRISE ===
Raison sociale : {ent['raison_sociale']}
SIRET : {ent['siret']}
TVA : Exoneree (article 261-4-4 du CGI)
{reco_bloc}{dce_bloc}{decp_bloc}

=== STRUCTURE ===

# BPU - Bordereau de Prix Unitaires
Tableau : | N | Designation | Unite | Prix unitaire HT | Observations |
Lignes : formation presentiel, distanciel, preparation, suivi, coaching, e-learning, certification RS6776, diagnostic/audit, frais deplacement

# DPGF - Decomposition du Prix Global Forfaitaire
Tableau : | N | Phase / Prestation | Quantite | Unite | PU HT | Total HT |
Phases : preparatoire, formation, suivi, e-learning, option certification

Total HT + note TVA exoneree + "Prix fermes et non revisables"

=== REGLES ===
- Utilise les tarifs de la RECOMMANDATION DE PRIX si disponible, sinon la grille tarifaire
- Viser 80-90% du budget si connu
- Certification RS6776 en OPTION separee
- Format Markdown avec tableaux"""

    resultat = _appel_claude(prompt, max_tokens=4000)

    # NE PAS ajouter le commentaire strategie dans le document client
    # (information interne visible lors de la generation dans les logs)
    if strategie_commentaire:
        logger.info(f"BPU strategie: {strategie_commentaire.strip()}")

    return resultat


def _generer_planning(ao: dict) -> str:
    """Genere le planning previsionnel."""
    prompt = f"""Genere un PLANNING PREVISIONNEL detaille pour cet appel d'offres.

=== APPEL D'OFFRES ===
{_bloc_infos_ao(ao)}

=== STRUCTURE ===
Genere un planning en format tableau Markdown avec ces colonnes :
| Phase | Description | Semaines | Livrables |

REGLES DE FORMAT STRICTES :
- Dans les cellules du tableau, separer les elements par des virgules (PAS de <br>, PAS de bullet points)
- Pas de balises HTML dans le markdown
- Pas de diagramme de Gantt en texte/ASCII (un visuel HTML sera genere automatiquement)

Phases typiques :
1. Phase preparatoire : diagnostic, personnalisation contenus (S1-S2)
2. Deploiement / Formation : sessions par groupes (S3-Sx)
3. Suivi post-formation : hotline, coaching (Sx-Sy)
4. Bilan et evaluation : rapport final, recommandations (derniere semaine)

Adapter le nombre de semaines a la duree du marche (maximum 52 semaines meme pour les marches pluriannuels, detailler la premiere annee).

Format Markdown."""

    return _appel_claude(prompt, max_tokens=3000, modele=MODELE_LEGER)


def _generer_cv_formateurs(ao: dict) -> str:
    """Genere les fiches CV des formateurs pertinents (template, sans appel API)."""

    # --- 1. Extraire les mots-cles de l'AO (titre + description) ---
    texte_ao = (ao.get('titre', '') + ' ' + ao.get('description', '')).lower()
    # Nettoyer : retirer accents simples, ponctuation
    for c in ".,;:!?()[]{}\"'-/\\":
        texte_ao = texte_ao.replace(c, ' ')
    mots_ao = set(texte_ao.split())
    # Retirer les mots vides courants
    mots_vides = {'de', 'du', 'des', 'le', 'la', 'les', 'un', 'une', 'et', 'en',
                  'a', 'au', 'aux', 'pour', 'par', 'sur', 'dans', 'avec', 'qui',
                  'que', 'est', 'sont', 'ce', 'cette', 'ces', 'ou', 'son', 'sa',
                  'ses', 'se', 'ne', 'pas', 'plus', 'tout', 'tous', 'etre', 'avoir',
                  'd', 'l', 'n', 'qu', 'il', 'elle', 'on', 'nous', 'vous', 'leur',
                  'leurs', 'y', 'si', 'mais', 'donc', 'ni', 'car', 'entre', 'comme',
                  'the', 'of', 'and', 'for', 'to', 'in', 'marche', 'public', 'lot',
                  'accord', 'cadre', 'prestations', 'services'}
    mots_ao -= mots_vides

    # --- 2. Scorer chaque formateur par correspondance mots-cles ---
    scores = []
    for f in ENTREPRISE["formateurs"]:
        spec_txt = (f.get('specialites', '') + ' ' + f.get('role', '') + ' ' + f.get('experience', '')).lower()
        for c in ".,;:!?()[]{}\"'-/\\":
            spec_txt = spec_txt.replace(c, ' ')
        mots_formateur = set(spec_txt.split()) - mots_vides
        matches = mots_ao & mots_formateur
        scores.append((len(matches), matches, f))

    # Trier par score decroissant
    scores.sort(key=lambda x: x[0], reverse=True)

    # Toujours inclure Mickael Bertolla en premier (formateur principal)
    selectionnes = []
    reste = []
    for score, matches, f in scores:
        if f['nom'] == 'Mickael Bertolla':
            selectionnes.insert(0, (score, matches, f))
        else:
            reste.append((score, matches, f))

    # Ajouter 1-3 autres formateurs (ceux avec le meilleur score, min 1 match)
    nb_autres = min(3, len([r for r in reste if r[0] > 0]))
    nb_autres = max(1, nb_autres)  # Au moins 1 autre formateur
    for i in range(min(nb_autres, len(reste))):
        selectionnes.append(reste[i])

    # --- 3. Generer le markdown ---
    titre_ao = ao.get('titre', 'cet appel d\'offres')
    md = f"""# FICHES CV - EQUIPE PEDAGOGIQUE MOBILISEE

**Appel d'offres** : {titre_ao}
**Candidat** : {ENTREPRISE['nom']} ({ENTREPRISE['raison_sociale']})
**Nombre de formateurs mobilises** : {len(selectionnes)}

---

"""

    for score, matches, f in selectionnes:
        # Construire la liste de mots-cles matches pour la pertinence
        if matches:
            mots_pertinents = ", ".join(sorted(matches)[:8])
        else:
            mots_pertinents = "direction pedagogique et coordination generale"

        md += f"""## {f['nom']} - {f['role']}

### Formation et diplomes
{f['formation']}

### Specialites et competences cles
{f['specialites']}

### Experience professionnelle
{f['experience']}

### Pertinence pour cet AO
{f['nom']} est mobilise(e) sur cette mission en raison de son expertise en lien direct avec les besoins identifies : {mots_pertinents}. Son profil repond aux exigences de l'appel d'offres par ses competences en {f['specialites'].split(',')[0].strip()}.

---

"""

    md += f"""## Organisation de l'equipe

L'equipe proposee est coordonnee par **{ENTREPRISE['formateurs'][0]['nom']}** ({ENTREPRISE['formateurs'][0]['role']}), garant de la qualite pedagogique et de la coherence de l'intervention.

Chaque formateur dispose d'une expertise complementaire permettant de couvrir l'ensemble des besoins exprimes dans le cahier des charges.

**Certifications de l'organisme** : {', '.join(ENTREPRISE.get('certifications', []))}
"""

    return md


def _generer_dc1_dc2(ao: dict) -> str:
    """Genere le DC1/DC2 pre-rempli (template, pas d'appel API)."""
    ent = ENTREPRISE
    date_jour = datetime.now().strftime("%d/%m/%Y")

    return f"""# DC1 - LETTRE DE CANDIDATURE

## Identification du pouvoir adjudicateur
- Acheteur : {ao.get('acheteur', 'Non precise')}
- Objet du marche : {ao.get('titre', 'Non precise')}
- Reference : {ao.get('id', '')}

## Identification du candidat
- Denomination : {ent['raison_sociale']} (nom commercial : {ent['nom']})
- Forme juridique : {ent['forme_juridique']}
- SIRET : {ent['siret']}
- NDA : {ent['nda']}
- Adresse : {ent['adresse']}
- Representant legal : {ent['representant']}
- Telephone : {ent['telephone']}
- Email : {ent['email']}
- Site web : {ent['site_web']}

## Objet de la candidature
Le candidat presente sa candidature pour le marche designe ci-dessus.
Le candidat se presente seul (candidature individuelle).

## Attestation sur l'honneur
Le candidat declare sur l'honneur :
- Ne pas tomber sous le coup des interdictions de soumissionner prevues aux articles L.2141-1 a L.2141-5 et L.2141-7 a L.2141-11 du Code de la commande publique
- Etre en regle au regard des articles L.5212-1 a L.5212-11 du code du travail (obligation d'emploi des travailleurs handicapes)

Date : {date_jour}
Signature : {ent['representant']}

---

# DC2 - DECLARATION DU CANDIDAT INDIVIDUEL

## 1. Identification du candidat
- Denomination sociale : {ent['raison_sociale']}
- Nom commercial : {ent['nom']}
- SIRET : {ent['siret']}
- Code APE/NAF : 8559A (Formation continue d'adultes)
- NDA : {ent['nda']}
- Forme juridique : {ent['forme_juridique']}
- Date de creation : 2023
- Adresse : {ent['adresse']}

## 2. Representant habilite
- Nom : Mickael Bertolla
- Qualite : President
- Habilitation : directe (President de SASU)

## 3. Renseignements economiques et financiers
- Chiffre d'affaires global : 200 000+ EUR
- Chiffre d'affaires relatif aux prestations objet du marche : 200 000+ EUR
- Effectif moyen annuel : 1 salarie + reseau de 10 formateurs freelance

## 4. Capacites techniques et professionnelles
### Certifications
{chr(10).join(f'- {c}' for c in ent['certifications'])}

### References
- {ent['chiffres']}

### Moyens humains
- Equipe de 6 formateurs specialises en IA
- Reseau de 10+ formateurs freelance mobilisables
- Couverture nationale

## 5. Sous-traitance
Le candidat n'envisage pas de sous-traiter une partie du marche.

Date : {date_jour}
Signature : {ent['representant']}
"""


def _generer_analyse_gonogo(ao: dict, gng_result: dict) -> str:
    """Genere le document d'analyse Go/No-Go (pas d'appel API)."""
    decision = gng_result.get("decision", "N/A")
    score = gng_result.get("score", 0)
    raison = gng_result.get("raison", "")

    criteres_txt = ""
    for c in gng_result.get("criteres", []):
        criteres_txt += f"- {c['nom']}: {c['score']}/{c['max']} - {c.get('commentaire', '')}\n"

    atouts_txt = "\n".join(f"- {a}" for a in gng_result.get("atouts", [])) or "- Aucun atout majeur detecte"
    risques_txt = "\n".join(f"- {r}" for r in gng_result.get("risques", [])) or "- Aucun risque majeur detecte"

    return f"""# ANALYSE GO / NO-GO

## Decision : {decision}
**Score global : {score}/100**
{raison}

## Appel d'offres
- Titre : {ao.get('titre', 'N/A')}
- Acheteur : {ao.get('acheteur', 'N/A')}
- Budget : {ao.get('budget_estime', 'Non precise')} EUR
- Date limite : {ao.get('date_limite', 'Non precisee')}

## Criteres evalues
{criteres_txt}

## Atouts
{atouts_txt}

## Risques / Points de vigilance
{risques_txt}

## Recommandation
{"Dossier genere automatiquement. Relire le memoire technique et adapter avant soumission." if decision == "GO" else "Analyser le DCE en detail avant de poursuivre." if decision == "A EVALUER" else "AO non pertinent pour Almera."}

---
*Analyse generee automatiquement le {datetime.now().strftime('%d/%m/%Y a %H:%M')}*
"""


def _generer_programme_formation(ao: dict, dce_texte: str = "") -> str:
    """Genere le programme de formation detaille en selectionnant dans le catalogue."""
    # Charger le catalogue YAML
    catalogue_txt = _charger_catalogue()

    dce_bloc = ""
    if dce_texte:
        dce_bloc = f"""

=== EXTRAIT DU DCE ===
Utilise ces extraits pour adapter le programme aux exigences exactes du cahier
des charges (durees, publics, objectifs, modalites demandees) :
{dce_texte[:8000]}
==="""

    prompt = f"""Tu es un ingenieur pedagogique expert en formation professionnelle.
Genere un PROGRAMME DE FORMATION DETAILLE pour cet appel d'offres.

=== APPEL D'OFFRES ===
{_bloc_infos_ao(ao)}

=== CATALOGUE DE FORMATIONS ALMERA (a utiliser comme base) ===
{catalogue_txt}
{dce_bloc}

=== INSTRUCTIONS ===
1. Selectionne les 1 a 4 formations du catalogue les plus pertinentes pour cet AO
2. Pour CHAQUE formation selectionnee, genere une fiche programme complete avec :

### Intitule de la formation
- **Objectifs pedagogiques** (verbes d'action : maitriser, savoir utiliser, etre capable de...)
- **Public vise** : (adapte a l'acheteur)
- **Prerequis** : (issus du catalogue)
- **Duree** : X heures / X jours
- **Modalites** : Presentiel / Distanciel / Mixte
- **Nombre de participants** : 4 a 12 par session (adapter selon l'AO)

### Deroulement detaille
Pour chaque module :
| Horaire | Module | Contenu detaille | Methode pedagogique |

### Methodes pedagogiques
- Apports theoriques (max 30%)
- Demonstrations en direct
- Ateliers pratiques individuels et en groupe (min 60%)
- Etudes de cas metier adaptes au contexte de l'acheteur
- Quiz et evaluation des acquis

### Modalites d'evaluation
- Evaluation diagnostique en debut de formation (positionnement)
- Evaluations formatives tout au long de la formation (quiz, exercices)
- Evaluation sommative en fin de formation (cas pratique + QCM)
- Attestation de fin de formation
- Certification RS6776 (si applicable)

### Supports fournis
- Support de cours PDF
- Document de 250 prompts IA
- Acces a la plateforme e-learning pendant 3 mois
- Fiches memo par outil

3. N'invente PAS de formations qui n'existent pas dans le catalogue
4. Adapte le vocabulaire et les exemples au contexte de l'acheteur
5. Format Markdown avec titres ## et ### et tableaux
6. INTERDIT : pas de balises HTML (<br>, <strong>, etc.) dans le markdown. Separer les elements dans les cellules de tableau par des virgules.
7. Utiliser la troisieme personne ou l'impersonnel (pas de "j'ai selectionne", mais "sont proposees")"""

    return _appel_claude(prompt, max_tokens=6000)


def _charger_catalogue() -> str:
    """Charge un resume du catalogue de formations."""
    # Chercher le catalogue : d'abord dans dashboard/, sinon dans ao_hunter/
    catalogue_path = DASHBOARD_DIR / "catalogue_formations.yaml"
    if not catalogue_path.exists():
        catalogue_path = DASHBOARD_DIR.parent / "catalogue_formations.yaml"
    if not catalogue_path.exists():
        # Fallback : catalogue integre
        return """Formations disponibles :
- Formation IA Generative (1-2 jours) : ChatGPT, Claude, Midjourney, Flux
- ChatGPT Niveau 1 (1 jour) : Prise en main, prompt engineering
- ChatGPT Niveau 2 (1 jour) : Fonctionnalites avancees, GPTs, plugins
- Microsoft Copilot (1 jour) : Integration Office 365
- Midjourney (1-2 jours) : Creation d'images par IA
- SEO/GSO/GEO (1-2 jours) : Referencement et IA
- Agents IA & Automatisation (1-2 jours) : Make, n8n, agents
- Mistral & IA open source (1 jour) : Modeles francais
- Formation sur mesure : Programmes adaptes aux besoins"""
    try:
        import yaml
        data = yaml.safe_load(catalogue_path.read_text(encoding="utf-8"))
        formations = data.get("formations", [])
        # Resume compact pour le prompt
        lignes = []
        for f in formations[:15]:  # Max 15 pour ne pas exploser le prompt
            lignes.append(f"- {f.get('titre', '')} ({f.get('duree_jours', '?')}j / {f.get('duree_heures', '?')}h)")
            if f.get('objectifs'):
                for obj in f['objectifs'][:3]:
                    lignes.append(f"  * {obj}")
            if f.get('programme'):
                for mod in f['programme'][:4]:
                    lignes.append(f"  - Module: {mod.get('module', '')} ({mod.get('duree', '')})")
        return "\n".join(lignes)
    except Exception:
        return "Catalogue non disponible - generer un programme adapte au contexte de l'AO"


def _generer_references_clients(ao: dict) -> str:
    """Genere les fiches references clients detaillees (template)."""
    ent = ENTREPRISE

    # References par secteur (du config.yaml)
    references = [
        {"client": "Havas", "secteur": "Communication et publicite", "mission": "Formation IA generative pour les equipes creatives et strategiques", "montant": "15 000 - 30 000 EUR HT", "periode": "2023-2025", "nb_personnes": "80+", "contact": "Direction de la transformation digitale"},
        {"client": "Eiffage", "secteur": "BTP / Construction", "mission": "Formation IA pour les cadres dirigeants et chefs de projet", "montant": "20 000 - 40 000 EUR HT", "periode": "2023-2025", "nb_personnes": "60+", "contact": "Direction des Ressources Humaines"},
        {"client": "Carrefour", "secteur": "Grande distribution", "mission": "Acculturation IA et ChatGPT pour les equipes marketing et supply chain", "montant": "15 000 - 25 000 EUR HT", "periode": "2023-2025", "nb_personnes": "50+", "contact": "Direction Innovation"},
        {"client": "Orange", "secteur": "Telecommunications", "mission": "Formation IA generative et prompt engineering pour les equipes techniques", "montant": "20 000 - 35 000 EUR HT", "periode": "2024-2025", "nb_personnes": "40+", "contact": "Orange Campus"},
        {"client": "Caisse des Depots", "secteur": "Institution financiere publique", "mission": "Formation IA pour les agents et cadres de la CDC", "montant": "25 000 - 50 000 EUR HT", "periode": "2024-2025", "nb_personnes": "100+", "contact": "Direction de la transformation numerique"},
        {"client": "Eli Lilly", "secteur": "Pharmaceutique", "mission": "Formation IA generative pour les equipes R&D et marketing", "montant": "10 000 - 20 000 EUR HT", "periode": "2024-2025", "nb_personnes": "30+", "contact": "Direction Scientifique"},
        {"client": "3DS (Dassault Systemes)", "secteur": "Technologie / Ingenierie", "mission": "Formation IA et automatisation pour les ingenieurs", "montant": "15 000 - 30 000 EUR HT", "periode": "2024-2025", "nb_personnes": "40+", "contact": "Direction R&D"},
        {"client": "Action Logement", "secteur": "Logement social", "mission": "Acculturation IA et transformation digitale", "montant": "10 000 - 20 000 EUR HT", "periode": "2024-2025", "nb_personnes": "50+", "contact": "Direction Generale"},
        {"client": "CCI", "secteur": "Chambre de commerce", "mission": "Formation IA pour les conseillers et equipes d'accompagnement", "montant": "15 000 - 25 000 EUR HT", "periode": "2023-2025", "nb_personnes": "60+", "contact": "Direction Formation"},
    ]

    # Selectionner les 5 plus pertinentes (prioriser secteur public si AO public)
    texte_ao = (ao.get("titre", "") + " " + ao.get("description", "")).lower()
    scored = []
    for ref in references:
        score = 0
        secteur_lower = ref["secteur"].lower()
        if "public" in texte_ao or "collectivit" in texte_ao or "administration" in texte_ao:
            if ref["client"] in ("Caisse des Depots", "CCI", "Action Logement"):
                score += 10
        for mot in secteur_lower.split():
            if mot in texte_ao:
                score += 3
        scored.append((score, ref))
    scored.sort(key=lambda x: x[0], reverse=True)
    top_refs = [r for _, r in scored[:5]]

    fiches = f"""# REFERENCES CLIENTS DETAILLEES

## Almera ({ent['nom']}) - {ent['raison_sociale']}
**{ent['chiffres']}**

---
"""
    for i, ref in enumerate(top_refs, 1):
        fiches += f"""
## Reference {i} : {ref['client']}

| Element | Detail |
|---------|--------|
| **Client** | {ref['client']} |
| **Secteur** | {ref['secteur']} |
| **Objet de la mission** | {ref['mission']} |
| **Montant** | {ref['montant']} |
| **Periode** | {ref['periode']} |
| **Nombre de personnes formees** | {ref['nb_personnes']} |
| **Contact** | {ref['contact']} |
| **Certifications utilisees** | Qualiopi, RS6776 France Competences |
| **Satisfaction** | 4.9/5 (moyenne Google) |

---
"""

    fiches += f"""
## Synthese

| Indicateur | Valeur |
|------------|--------|
| Nombre total de personnes formees | 2 000+ |
| Nombre d'entreprises clientes | 50+ |
| Note de satisfaction moyenne | 4.9/5 (Google) |
| Taux de recommandation | > 95% |
| Anciennete | Depuis 2022 |

*References disponibles sur demande. Attestations de bonne execution fournies sur demande.*

---
*Document genere le {datetime.now().strftime('%d/%m/%Y')}*
"""
    return fiches


def _generer_dpgf_excel(ao: dict, dossier_path) -> str | None:
    """Genere le DPGF au format Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        logger.warning("openpyxl non installe - DPGF Excel non genere")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "BPU - DPGF"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    money_fmt = '#,##0.00 €'

    ent = ENTREPRISE
    budget = ao.get("budget_estime")

    # --- En-tete ---
    ws.merge_cells("A1:F1")
    ws["A1"] = f"BORDEREAU DE PRIX UNITAIRES (BPU) - {ent['raison_sociale']} ({ent['nom']})"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Marche : {ao.get('titre', 'Non precise')}"
    ws["A2"].font = Font(name="Calibri", size=11)
    ws.merge_cells("A3:F3")
    ws["A3"] = f"Acheteur : {ao.get('acheteur', 'Non precise')}"
    ws["A3"].font = Font(name="Calibri", size=11)
    ws.merge_cells("A4:F4")
    ws["A4"] = f"SIRET : {ent['siret']} | NDA : {ent['nda']}"

    # --- BPU ---
    row = 6
    headers_bpu = ["N°", "Designation de la prestation", "Unite", "Prix unitaire HT", "Observations"]
    for col, h in enumerate(headers_bpu, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    bpu_data = [
        ("1", "Journee de formation en presentiel (groupe 12 max)", "Jour", 1500, ""),
        ("2", "Journee de formation en distanciel (groupe 12 max)", "Jour", 1200, ""),
        ("3", "Journee de preparation / personnalisation contenus", "Jour", 800, ""),
        ("4", "Heure de suivi post-formation / hotline", "Heure", 150, ""),
        ("5", "Journee de coaching individuel", "Jour", 1800, ""),
        ("6", "Acces plateforme e-learning (par stagiaire)", "Stagiaire", 200, "3 mois inclus"),
        ("7", "Certification RS6776 (par stagiaire)", "Stagiaire", 500, "Option"),
        ("8", "Journee de diagnostic / audit des besoins", "Jour", 1500, ""),
        ("9", "Frais de deplacement (hors Ile-de-France)", "Forfait", 250, "Si applicable"),
    ]

    for data_row in bpu_data:
        row += 1
        for col, val in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
            if col == 4 and isinstance(val, (int, float)):
                cell.number_format = money_fmt

    # --- DPGF ---
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="DECOMPOSITION DU PRIX GLOBAL FORFAITAIRE (DPGF)").font = Font(name="Calibri", bold=True, size=13, color="1E3A5F")

    row += 1
    headers_dpgf = ["N°", "Phase / Prestation", "Quantite", "Unite", "PU HT", "Total HT"]
    for col, h in enumerate(headers_dpgf, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Adapter les quantites au budget
    if budget and budget > 0:
        nb_jours_form = max(2, min(10, int(budget * 0.5 / 1500)))
        nb_jours_prep = max(1, nb_jours_form // 3)
        nb_stagiaires = max(6, min(50, int(budget * 0.1 / 200)))
    else:
        nb_jours_form = 4
        nb_jours_prep = 2
        nb_stagiaires = 12

    dpgf_data = [
        ("1", "Phase preparatoire (diagnostic, personnalisation)", nb_jours_prep, "Jour", 800, nb_jours_prep * 800),
        ("2", "Formation presentiel (sessions par groupes)", nb_jours_form, "Jour", 1500, nb_jours_form * 1500),
        ("3", "Suivi post-formation (hotline, coaching)", nb_jours_form * 2, "Heure", 150, nb_jours_form * 2 * 150),
        ("4", "Acces plateforme e-learning", nb_stagiaires, "Stagiaire", 200, nb_stagiaires * 200),
    ]

    total_ht = 0
    for data_row in dpgf_data:
        row += 1
        for col, val in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
            if col in (5, 6) and isinstance(val, (int, float)):
                cell.number_format = money_fmt
        total_ht += data_row[5]

    # Option certification
    row += 1
    cert_total = nb_stagiaires * 500
    for col, val in enumerate(("OPT", "OPTION - Certification RS6776", nb_stagiaires, "Stagiaire", 500, cert_total), 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = border
        cell.font = Font(name="Calibri", italic=True)
        if col in (5, 6) and isinstance(val, (int, float)):
            cell.number_format = money_fmt

    # Totaux
    row += 1
    ws.cell(row=row, column=5, value="Sous-total HT").font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_ht).font = Font(bold=True)
    ws.cell(row=row, column=6).number_format = money_fmt
    ws.cell(row=row, column=6).border = border

    row += 1
    ws.cell(row=row, column=5, value="TVA").font = Font(bold=True)
    ws.cell(row=row, column=6, value="Exoneree (art. 261-4-4 CGI)")

    row += 1
    ws.cell(row=row, column=5, value="TOTAL TTC").font = Font(bold=True, size=12)
    ws.cell(row=row, column=6, value=total_ht).font = Font(bold=True, size=12)
    ws.cell(row=row, column=6).number_format = money_fmt
    ws.cell(row=row, column=6).border = border

    row += 2
    ws.cell(row=row, column=1, value="Prix fermes et non revisables pour la duree du marche.")
    ws.cell(row=row, column=1).font = Font(italic=True)

    # Largeurs de colonnes
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Sauvegarder
    excel_path = dossier_path / "04_bpu_dpgf.xlsx"
    wb.save(str(excel_path))
    return "04_bpu_dpgf.xlsx"


def _generer_acte_engagement(ao: dict) -> str:
    """Genere l'acte d'engagement pre-rempli (template)."""
    ent = ENTREPRISE
    date_jour = datetime.now().strftime("%d/%m/%Y")
    date_limite = ao.get("date_limite", "Non precisee")
    if "T" in str(date_limite):
        date_limite = date_limite.split("T")[0]

    budget = ao.get("budget_estime")
    budget_txt = f"{budget:,.0f} EUR HT" if budget else "[A completer selon le BPU/DPGF]"

    return f"""# ACTE D'ENGAGEMENT

## A. OBJET DU MARCHE

### Identification du pouvoir adjudicateur
- **Acheteur** : {ao.get('acheteur', '[Nom de l acheteur]')}
- **Adresse** : Voir avis de marche

### Objet du marche
- **Intitule** : {ao.get('titre', '[Objet du marche]')}
- **Reference** : {ao.get('id', '[Reference du marche]')}
- **Type** : {ao.get('type_marche', 'Services')}
- **Procedure** : {ao.get('type_procedure', 'Procedure adaptee')}

### Duree du marche
- **Duree** : {ao.get('duree_mois', '[A preciser]')} mois
- **Date previsionnelle de debut** : A preciser par l'acheteur

---

## B. IDENTIFICATION DU CANDIDAT

| Element | Valeur |
|---------|--------|
| Denomination sociale | {ent['raison_sociale']} |
| Nom commercial | {ent['nom']} |
| Forme juridique | {ent['forme_juridique']} |
| SIRET | {ent['siret']} |
| NDA | {ent['nda']} |
| Code APE/NAF | 8559A |
| Adresse du siege | {ent['adresse']} |
| Representant legal | {ent['representant']} |
| Telephone | {ent['telephone']} |
| Email | {ent['email']} |
| Site web | {ent['site_web']} |

---

## C. ENGAGEMENTS DU CANDIDAT

Le soussigne, **Mickael Bertolla**, agissant en qualite de **President**
de la societe **{ent['raison_sociale']}** ({ent['nom']}),

Apres avoir pris connaissance de l'ensemble des pieces constitutives du
marche et de ses annexes,

**S'engage**, sans reserve, conformement aux stipulations des documents
du marche, a executer les prestations objet du present marche aux
conditions suivantes :

### Prix du marche
- **Montant total HT** : {budget_txt}
- **TVA** : Exoneree (article 261-4-4 du Code General des Impots -
  organismes de formation)
- **Montant total TTC** : Egal au montant HT
- **Caractere des prix** : Fermes et non revisables

### Delai de validite de l'offre
L'offre est valable **120 jours** a compter de la date limite de remise
des offres ({date_limite}).

### Domiciliation bancaire
- **Titulaire du compte** : {ent['raison_sociale']}
- **Banque** : Voir RIB joint au dossier

---

## D. DECLARATIONS

Le candidat declare :
- Avoir pris connaissance du reglement de consultation, du CCAP, du CCTP
  et de l'ensemble des pieces du DCE
- Ne pas tomber sous le coup des interdictions de soumissionner prevues
  aux articles L.2141-1 a L.2141-5 du Code de la commande publique
- Etre en regle au regard de ses obligations fiscales et sociales
- Etre en regle au regard de l'obligation d'emploi des travailleurs
  handicapes (articles L.5212-1 a L.5212-11 du Code du travail)
- Ne pas avoir fait l'objet d'une condamnation inscrite au bulletin n°2
  du casier judiciaire pour les infractions mentionnees a l'article
  L.2141-1 du Code de la commande publique

---

## E. SIGNATURE

Fait a **Paris**, le **{date_jour}**

Le representant legal du candidat,

**{ent['representant']}**

*(Signature et cachet de l'entreprise)*

---
*Note : Si l'acheteur fournit un formulaire d'Acte d'Engagement dans le DCE,
utiliser ce formulaire en priorite et reporter les informations ci-dessus.*

*Document genere le {date_jour}*
"""


def _generer_dume(ao: dict) -> str:
    """Genere le Document Unique de Marche Europeen (DUME) pre-rempli."""
    ent = ENTREPRISE
    date_jour = datetime.now().strftime("%d/%m/%Y")

    return f"""# DOCUMENT UNIQUE DE MARCHE EUROPEEN (DUME)
## Reponse au marche : {ao.get('titre', 'Non precise')}

---

## PARTIE I : IDENTIFICATION DE LA PROCEDURE

| Element | Valeur |
|---------|--------|
| Intitule du marche | {ao.get('titre', 'Non precise')} |
| Acheteur | {ao.get('acheteur', 'Non precise')} |
| Reference / numero de l'avis | {ao.get('id', '')} |
| Type de procedure | {ao.get('type_procedure', 'Non precise')} |

---

## PARTIE II : INFORMATIONS CONCERNANT L'OPERATEUR ECONOMIQUE

### A. Informations generales

| Element | Valeur |
|---------|--------|
| Denomination officielle | {ent['raison_sociale']} |
| Nom commercial | {ent['nom']} |
| Numero d'identification (SIRET) | {ent['siret']} |
| Numero de TVA intracommunautaire | FR 12 989004551 |
| Adresse postale | {ent['adresse']} |
| Personne de contact | {ent['representant']} |
| Telephone | {ent['telephone']} |
| Courriel | {ent['email']} |
| Adresse internet | https://{ent['site_web']} |
| Forme juridique | {ent['forme_juridique']} |
| Code NACE / APE | 8559A - Formation continue d'adultes |
| Taille de l'entreprise | Micro-entreprise (< 10 salaries) |
| NDA (formation) | {ent['nda']} |

### B. Representant(s) de l'operateur economique
- **Nom** : Mickael Bertolla
- **Qualite** : President
- **Date de naissance** : Communiquee sur demande
- **Habilitation** : Representant legal (President de SASU)

### C. Sous-traitance
L'operateur economique n'a pas l'intention de sous-traiter une partie du marche.

---

## PARTIE III : CRITERES D'EXCLUSION

### A. Motifs lies a des condamnations penales
Le candidat declare sur l'honneur qu'il n'a pas fait l'objet de
condamnations pour les infractions visees aux articles L.2141-1 a
L.2141-5 du Code de la commande publique.

### B. Motifs lies au paiement d'impots et de cotisations
Le candidat declare sur l'honneur etre en regle de ses obligations
fiscales et sociales. Justificatifs joints au dossier :
- Attestation fiscale (DGFIP)
- Attestation de vigilance URSSAF

### C. Motifs lies a l'insolvabilite
Le candidat declare ne pas se trouver en situation de redressement
judiciaire, liquidation ou faillite.

---

## PARTIE IV : CRITERES DE SELECTION

### A. Aptitude a exercer l'activite professionnelle

| Element | Valeur |
|---------|--------|
| Inscription registre du commerce | Oui (Kbis joint) |
| Certification Qualiopi | Oui (certificat joint, renouvellement en cours) |
| Certification RS6776 | Oui (France Competences - IA generative) |
| Activateur France Num | Oui (certificat joint) |
| NDA formation | {ent['nda']} |

### B. Capacite economique et financiere

| Element | Valeur |
|---------|--------|
| Chiffre d'affaires annuel | 200 000+ EUR |
| CA relatif au marche | 200 000+ EUR |
| Assurance RC professionnelle | En cours d'obtention |

### C. Capacites techniques et professionnelles

| Element | Valeur |
|---------|--------|
| Nombre de personnes formees | 2 000+ |
| Nombre d'entreprises clientes | 50+ |
| Note de satisfaction | 4.9/5 (Google) |
| Effectif | 1 salarie + reseau 10+ formateurs freelance |
| Couverture geographique | France entiere |
| Certifications qualite | Qualiopi, RS6776, France Num |

### D. References principales
- Havas, Eiffage, Carrefour, Orange (grands groupes)
- Caisse des Depots, CCI, Action Logement (secteur public)
- 3DS (Dassault Systemes), Eli Lilly (international)

---

## PARTIE V : REDUCTION DU NOMBRE DE CANDIDATS QUALIFIES

Sans objet (candidature unique).

---

## PARTIE VI : DECLARATIONS FINALES

Le soussigne declare formellement que les renseignements fournis dans les
parties II a V ci-dessus sont exacts et complets et qu'ils ont ete
etablis en pleine connaissance de cause.

Le soussigne declare formellement etre en mesure de produire, sur
demande et sans delai, les certificats et autres formes de preuves
documentaires visees.

Fait a Paris, le {date_jour}

**{ent['representant']}**

---
*Note : Pour les marches europeens, utiliser le formulaire DUME
electronique (espd.uzpe.gov.fr) en priorite. Ce document sert de
base pour pre-remplir le formulaire officiel.*

*Document genere le {date_jour}*
"""


def _generer_moyens_techniques(ao: dict) -> str:
    """Genere la fiche moyens techniques et materiels."""
    ent = ENTREPRISE

    return f"""# MOYENS TECHNIQUES ET MATERIELS

## Almera ({ent['raison_sociale']}) - Moyens mis a disposition

---

## 1. Outils IA maitrises et utilises en formation

### IA Generative de Texte
| Outil | Usage | Licence |
|-------|-------|---------|
| ChatGPT (OpenAI) | Formation, generation de contenu, analyse | Teams / Enterprise |
| Claude (Anthropic) | Formation, analyse documentaire, code | Pro |
| Mistral AI | Formation, IA souveraine francaise | API |
| Gemini (Google) | Formation, integration workspace | Enterprise |
| Perplexity AI | Recherche augmentee par IA | Pro |
| Copilot (Microsoft) | Integration Office 365 | Business |

### IA Generative d'Image
| Outil | Usage |
|-------|-------|
| Midjourney | Creation d'images professionnelles |
| DALL-E 3 (ChatGPT) | Generation d'images integree |
| Krea AI | Design et creation visuelle |
| Flux (Black Forest Labs) | Generation d'images open source |
| Leonardo AI | Design et assets visuels |

### IA Audio & Video
| Outil | Usage |
|-------|-------|
| ElevenLabs | Synthese vocale, doublage |
| HeyGen | Avatars video IA |
| Runway ML | Generation et edition video |

### Automatisation & Agents IA
| Outil | Usage |
|-------|-------|
| Make (Integromat) | Automatisation de workflows |
| n8n | Automatisation open source |
| Zapier | Integration d'applications |
| Voiceflow | Chatbots et agents conversationnels |

---

## 2. Plateformes pedagogiques

| Plateforme | Usage |
|------------|-------|
| Plateforme e-learning Almera | Acces post-formation (3 mois), modules de revision |
| Google Workspace | Supports de cours, partage collaboratif |
| Notion | Documentation et ressources partagees |
| Zoom / Teams / Google Meet | Formations a distance |

---

## 3. Supports de formation

| Type | Description |
|------|-------------|
| Supports PDF | Presentations detaillees par module |
| Document 250 prompts | Guide de prompts IA par metier |
| Fiches memo | Aide-memoire par outil (ChatGPT, Copilot, Midjourney...) |
| Cahier d'exercices | Cas pratiques et ateliers |
| Acces replays | Enregistrements des sessions (si distanciel) |

---

## 4. Equipements

| Element | Detail |
|---------|--------|
| Postes formateur | Laptop + ecran externe + tablette graphique |
| Licences logicielles | Toutes licences Pro/Enterprise a jour |
| Connectivite | 4G/5G de secours en cas de probleme reseau |
| Videoprojecteur portable | En cas de besoin (formation sur site) |

---

## 5. Locaux

| Element | Detail |
|---------|--------|
| Formation sur site client | Almera se deplace dans les locaux de l'acheteur |
| Salles de formation partenaires | Possibilite de reservation a Paris et en regions |
| Formation a distance | Via Zoom, Teams ou Google Meet |

---

## 6. Accessibilite

Almera est sensible a l'accessibilite et peut adapter ses formations
pour les personnes en situation de handicap :
- Supports en format accessible
- Adaptation du rythme et des modalites
- Referent handicap : Mickael Bertolla (contact@almera.one)

---
*Document genere le {datetime.now().strftime('%d/%m/%Y')}*
"""


def _generer_checklist_soumission(ao: dict, fichiers_generes: list[str]) -> str:
    """Genere la checklist de soumission (pas d'appel API)."""
    date_limite = ao.get("date_limite", "Non precisee")
    if "T" in str(date_limite):
        date_limite = date_limite.split("T")[0]

    fichiers_txt = "\n".join(f"- [x] {f}" for f in fichiers_generes)

    return f"""# CHECKLIST DE SOUMISSION

## Appel d'offres
- Titre : {ao.get('titre', 'N/A')}
- Acheteur : {ao.get('acheteur', 'N/A')}
- Date limite : {date_limite}

## Documents generes automatiquement
{fichiers_txt}

## Pieces administratives a joindre depuis le dossier permanent
- [ ] Attestation sur l'honneur (01_attestation_honneur.pdf)
- [ ] Kbis < 3 mois (02_kbis.pdf - verifier la date)
- [ ] Attestation URSSAF de vigilance (03_attestation_urssaf_vigilance.pdf)
- [ ] Attestation fiscale DGFIP (04_attestation_fiscale.pdf)
- [ ] Assurance RC Pro (05_assurance_rc_pro.pdf - EN ATTENTE)
- [ ] Certificat Qualiopi (06_certificat_qualiopi.pdf)
- [ ] RIB (08_rib.pdf)

## Verification avant envoi
- [ ] Relire le memoire technique (coherence, personnalisation)
- [ ] Verifier les prix du BPU/DPGF
- [ ] Verifier le planning (dates realistes)
- [ ] Completer le DC1 si formulaire officiel fourni
- [ ] Signer les documents necessaires
- [ ] Verifier le format de depot (plateforme, email, courrier)
- [ ] Deposer AVANT la date limite ({date_limite})

## Rappels
- Depot sur la plateforme de l'acheteur (verifier URL dans l'avis)
- Conserver l'accuse de depot
- Anticiper les problemes techniques (depot 24h avant la deadline)

---
*Checklist generee automatiquement le {datetime.now().strftime('%d/%m/%Y a %H:%M')}*
"""


# --- Fonction principale ---

def generer_dossier_complet(ao: dict, type_presta: str = "Formation", gng_result: dict = None, dce_texte: str = "") -> dict:
    """Genere un dossier de candidature COMPLET via l'API Claude.

    Args:
        ao: Dictionnaire de l'appel d'offres
        type_presta: Type de prestation detecte
        gng_result: Resultat du Go/No-Go (si deja calcule)
        dce_texte: Texte extrait du DCE (optionnel, enrichit les prompts)

    Returns:
        dict avec: success, dossier_nom, fichiers, nb_mots, erreur
    """
    if not API_KEY:
        return {"success": False, "erreur": "ANTHROPIC_API_KEY non configuree sur Render"}

    try:
        import httpx
    except ImportError:
        return {"success": False, "erreur": "httpx non installe"}

    # Creer le dossier
    clean_id = ao.get("id", "inconnu").replace("/", "_").replace("\\", "_")
    dossier_nom = f"AO_{clean_id}_{datetime.now():%Y%m%d}"
    dossier_path = DOSSIERS_DIR / dossier_nom
    dossier_path.mkdir(parents=True, exist_ok=True)

    fichiers_generes = []
    nb_mots_total = 0
    erreurs = []

    # Feature 2 : Personnalisation automatique par type d'acheteur
    personnalisation = None
    try:
        from personnalisation_acheteur import personnaliser
        # Passer les references du fichier _generer_references_clients
        refs_list = [
            {"client": "Havas", "secteur": "Communication et publicite", "mission": "Formation IA generative pour les equipes creatives et strategiques"},
            {"client": "Eiffage", "secteur": "BTP / Construction", "mission": "Formation IA pour les cadres dirigeants et chefs de projet"},
            {"client": "Carrefour", "secteur": "Grande distribution", "mission": "Acculturation IA et ChatGPT pour les equipes marketing et supply chain"},
            {"client": "Orange", "secteur": "Telecommunications", "mission": "Formation IA generative et prompt engineering pour les equipes techniques"},
            {"client": "Caisse des Depots", "secteur": "Institution financiere publique", "mission": "Formation IA pour les agents et cadres de la CDC"},
            {"client": "Eli Lilly", "secteur": "Pharmaceutique", "mission": "Formation IA generative pour les equipes R&D et marketing"},
            {"client": "3DS (Dassault Systemes)", "secteur": "Technologie / Ingenierie", "mission": "Formation IA et automatisation pour les ingenieurs"},
            {"client": "Action Logement", "secteur": "Logement social", "mission": "Acculturation IA et transformation digitale"},
            {"client": "CCI", "secteur": "Chambre de commerce", "mission": "Formation IA pour les conseillers et equipes d'accompagnement"},
        ]
        personnalisation = personnaliser(ao, references=refs_list)
        logger.info(f"Personnalisation: type_acheteur={personnalisation['type_acheteur']}")
    except Exception as e:
        logger.warning(f"Personnalisation acheteur non disponible: {e}")

    # Données DECP réelles (intelligence concurrentielle)
    decp = None
    try:
        from decp_data import rechercher_marches_similaires
        decp = rechercher_marches_similaires(ao)
        if decp and decp.get("nb_marches_trouves", 0) > 0:
            logger.info(f"DECP: {decp['nb_marches_trouves']} marches similaires trouves")
        else:
            logger.info("DECP: aucun marche similaire trouve")
    except Exception as e:
        logger.warning(f"DECP data non disponible: {e}")

    # 1. Analyse Go/No-Go (pas d'appel API)
    logger.info("1/7 - Analyse Go/No-Go...")
    try:
        if gng_result:
            analyse_txt = _generer_analyse_gonogo(ao, gng_result)
        else:
            from analyse_dce import go_no_go
            gng_result = go_no_go(ao)
            analyse_txt = _generer_analyse_gonogo(ao, gng_result)
        _sauvegarder(dossier_path, "01_analyse_go_no_go.md", analyse_txt)
        fichiers_generes.append("01_analyse_go_no_go.md")
    except Exception as e:
        erreurs.append(f"Analyse Go/No-Go: {e}")
        logger.error(f"Erreur analyse: {e}")

    # 2. Memoire technique (appel API principal)
    logger.info("2/7 - Memoire technique...")
    try:
        memoire = _generer_memoire(ao, type_presta, dce_texte=dce_texte, personnalisation=personnalisation, decp=decp)
        _sauvegarder(dossier_path, "02_memoire_technique.md", memoire)
        fichiers_generes.append("02_memoire_technique.md")
        nb_mots_total += len(memoire.split())
        logger.info(f"  Memoire: {len(memoire.split())} mots")
    except Exception as e:
        erreurs.append(f"Memoire technique: {e}")
        logger.error(f"Erreur memoire: {e}")

    # 3. Lettre de candidature (appel API)
    logger.info("3/7 - Lettre de candidature...")
    try:
        lettre = _generer_lettre(ao, personnalisation=personnalisation)
        _sauvegarder(dossier_path, "03_lettre_candidature.md", lettre)
        fichiers_generes.append("03_lettre_candidature.md")
        nb_mots_total += len(lettre.split())
    except Exception as e:
        erreurs.append(f"Lettre: {e}")
        logger.error(f"Erreur lettre: {e}")

    # 4. BPU / DPGF (appel API)
    logger.info("4/7 - BPU / DPGF...")
    try:
        bpu = _generer_bpu(ao, dce_texte=dce_texte, decp=decp)
        _sauvegarder(dossier_path, "04_bpu_dpgf.md", bpu)
        fichiers_generes.append("04_bpu_dpgf.md")
        nb_mots_total += len(bpu.split())
    except Exception as e:
        erreurs.append(f"BPU: {e}")
        logger.error(f"Erreur BPU: {e}")

    # 5. Planning (appel API leger)
    logger.info("5/7 - Planning...")
    try:
        planning = _generer_planning(ao)
        _sauvegarder(dossier_path, "05_planning_previsionnel.md", planning)
        fichiers_generes.append("05_planning_previsionnel.md")
        nb_mots_total += len(planning.split())
    except Exception as e:
        erreurs.append(f"Planning: {e}")
        logger.error(f"Erreur planning: {e}")

    # 6. CV Formateurs (appel API leger)
    logger.info("6/7 - CV Formateurs...")
    try:
        cv = _generer_cv_formateurs(ao)
        _sauvegarder(dossier_path, "06_cv_formateurs.md", cv)
        fichiers_generes.append("06_cv_formateurs.md")
        nb_mots_total += len(cv.split())
    except Exception as e:
        erreurs.append(f"CV: {e}")
        logger.error(f"Erreur CV: {e}")

    # 7. DC1 / DC2 (template, pas d'appel API)
    logger.info("7/13 - DC1 / DC2...")
    try:
        dc = _generer_dc1_dc2(ao)
        _sauvegarder(dossier_path, "07_dc1_dc2.md", dc)
        fichiers_generes.append("07_dc1_dc2.md")
        nb_mots_total += len(dc.split())
    except Exception as e:
        erreurs.append(f"DC1/DC2: {e}")
        logger.error(f"Erreur DC: {e}")

    # 8. Programme de formation detaille (appel API)
    logger.info("8/13 - Programme de formation detaille...")
    try:
        programme = _generer_programme_formation(ao, dce_texte=dce_texte)
        _sauvegarder(dossier_path, "08_programme_formation.md", programme)
        fichiers_generes.append("08_programme_formation.md")
        nb_mots_total += len(programme.split())
        logger.info(f"  Programme: {len(programme.split())} mots")
    except Exception as e:
        erreurs.append(f"Programme formation: {e}")
        logger.error(f"Erreur programme: {e}")

    # 9. References clients detaillees (template)
    logger.info("9/13 - References clients...")
    try:
        refs = _generer_references_clients(ao)
        _sauvegarder(dossier_path, "09_references_clients.md", refs)
        fichiers_generes.append("09_references_clients.md")
        nb_mots_total += len(refs.split())
    except Exception as e:
        erreurs.append(f"References: {e}")
        logger.error(f"Erreur references: {e}")

    # 10. DPGF en Excel (openpyxl)
    logger.info("10/13 - DPGF Excel...")
    try:
        excel_file = _generer_dpgf_excel(ao, dossier_path)
        if excel_file:
            fichiers_generes.append(excel_file)
            logger.info("  DPGF Excel genere")
    except Exception as e:
        erreurs.append(f"DPGF Excel: {e}")
        logger.error(f"Erreur DPGF Excel: {e}")

    # 11. Acte d'engagement pre-rempli (template)
    logger.info("11/13 - Acte d'engagement...")
    try:
        ae = _generer_acte_engagement(ao)
        _sauvegarder(dossier_path, "11_acte_engagement.md", ae)
        fichiers_generes.append("11_acte_engagement.md")
        nb_mots_total += len(ae.split())
    except Exception as e:
        erreurs.append(f"Acte engagement: {e}")
        logger.error(f"Erreur AE: {e}")

    # 12. DUME (template)
    logger.info("12/13 - DUME...")
    try:
        dume = _generer_dume(ao)
        _sauvegarder(dossier_path, "12_dume.md", dume)
        fichiers_generes.append("12_dume.md")
        nb_mots_total += len(dume.split())
    except Exception as e:
        erreurs.append(f"DUME: {e}")
        logger.error(f"Erreur DUME: {e}")

    # 13. Moyens techniques et materiels (template)
    logger.info("13/13 - Moyens techniques...")
    try:
        moyens = _generer_moyens_techniques(ao)
        _sauvegarder(dossier_path, "13_moyens_techniques.md", moyens)
        fichiers_generes.append("13_moyens_techniques.md")
        nb_mots_total += len(moyens.split())
    except Exception as e:
        erreurs.append(f"Moyens techniques: {e}")
        logger.error(f"Erreur moyens: {e}")

    # 14. Annexes visuelles HTML (planning Gantt, organigramme, radar, references, fiche synthese)
    logger.info("14/17 - Annexes visuelles...")
    try:
        from annexes_visuelles import generer_toutes_annexes
        annexes = generer_toutes_annexes(ao, dossier_path)
        fichiers_generes.extend(annexes)
        logger.info(f"  {len(annexes)} annexe(s) visuelle(s) generee(s)")
    except Exception as e:
        erreurs.append(f"Annexes visuelles: {e}")
        logger.error(f"Erreur annexes visuelles: {e}")

    # 15. Checklist (template, pas d'appel API)
    try:
        checklist = _generer_checklist_soumission(ao, fichiers_generes)
        _sauvegarder(dossier_path, "15_checklist_soumission.md", checklist)
        fichiers_generes.append("15_checklist_soumission.md")
    except Exception as e:
        erreurs.append(f"Checklist: {e}")

    # 16. Conversion des documents cles en DOCX
    try:
        _convertir_en_docx(dossier_path, ao, fichiers_generes)
    except Exception as e:
        erreurs.append(f"Conversion DOCX: {e}")
        logger.error(f"Erreur conversion DOCX globale: {e}")

    # 17. Pre-remplissage formulaires natifs du DCE (DC1/DC2/DUME PDF/DOCX)
    try:
        from formulaires_natifs import preremplir_formulaires
        # Chercher le dossier DCE (telecharge par dce_auto)
        clean_id = ao.get("id", "inconnu").replace("/", "_").replace("\\", "_")
        dossier_dce_candidat = DOSSIERS_DIR / f"DCE_{clean_id}"
        if not dossier_dce_candidat.exists():
            # Tenter aussi le dossier courant (les fichiers DCE sont parfois dans le meme dossier)
            dossier_dce_candidat = dossier_path
        formulaires_ok = preremplir_formulaires(dossier_dce_candidat, dossier_path)
        if formulaires_ok:
            fichiers_generes.extend(formulaires_ok)
            logger.info(f"  {len(formulaires_ok)} formulaire(s) pre-rempli(s)")
    except Exception as e:
        erreurs.append(f"Pre-remplissage formulaires: {e}")
        logger.error(f"Erreur pre-remplissage formulaires: {e}")

    # 18. Fiche AO en JSON
    fiche_path = dossier_path / "fiche_ao.json"
    fiche_path.write_text(json.dumps(ao, ensure_ascii=False, indent=2), encoding="utf-8")
    fichiers_generes.append("fiche_ao.json")

    # Feature 3 : Auto-review IA du dossier avant soumission
    review_result = None
    logger.info("Auto-review IA du dossier...")
    try:
        from auto_review import review_dossier

        # Charger tous les fichiers markdown generes
        fichiers_contenu = {}
        for f_nom in fichiers_generes:
            if f_nom.endswith(".md"):
                f_path = dossier_path / f_nom
                if f_path.exists():
                    fichiers_contenu[f_nom] = f_path.read_text(encoding="utf-8")

        criteres_pour_review = _extraire_criteres_attribution(ao, dce_texte)
        review_result = review_dossier(fichiers_contenu, ao, criteres_attribution=criteres_pour_review)

        # Sauvegarder le resultat dans review_auto.json
        review_path = dossier_path / "review_auto.json"
        review_path.write_text(json.dumps(review_result, ensure_ascii=False, indent=2), encoding="utf-8")
        fichiers_generes.append("review_auto.json")
        logger.info(f"Auto-review: score={review_result.get('score_qualite', '?')}/100, conforme={review_result.get('conforme', '?')}")
    except Exception as e:
        erreurs.append(f"Auto-review: {e}")
        logger.error(f"Erreur auto-review: {e}")

    # Feature : Verification conformite RC (si donnees RC disponibles)
    conformite_rc_result = None
    logger.info("Verification conformite RC...")
    try:
        from auto_review import verifier_conformite_rc
        from extraction_rc import extraire_rc

        # Chercher le dossier DCE pour en extraire le RC
        clean_id_rc = ao.get("id", "inconnu").replace("/", "_").replace("\\", "_")
        dossier_dce_rc = DOSSIERS_DIR / f"DCE_{clean_id_rc}"
        rc_data = {}
        if dossier_dce_rc.exists():
            rc_result = extraire_rc(dossier_dce_rc)
            # extraire_rc peut retourner un str (erreur) au lieu d'un dict
            if isinstance(rc_result, dict):
                rc_data = rc_result
            else:
                logger.warning(f"extraire_rc a retourne un str au lieu d'un dict: {str(rc_result)[:200]}")
        else:
            # Construire un rc_data minimal a partir des infos AO disponibles
            # Normaliser : certains champs AO sont des str au lieu de list
            def _as_list(v):
                if v is None:
                    return []
                if isinstance(v, list):
                    return v
                if isinstance(v, str):
                    return [line.strip() for line in v.split("\n") if line.strip()]
                return []

            rc_data = {
                "pieces_exigees": _as_list(ao.get("pieces_exigees")),
                "criteres_attribution": _as_list(ao.get("criteres_attribution")),
                "conditions_participation": _as_list(ao.get("conditions_participation")),
            }

        if rc_data.get("pieces_exigees") or rc_data.get("criteres_attribution"):
            # Charger les fichiers du dossier
            fichiers_pour_rc = {}
            for f_nom in fichiers_generes:
                if f_nom.endswith(".md"):
                    f_path = dossier_path / f_nom
                    if f_path.exists():
                        fichiers_pour_rc[f_nom] = f_path.read_text(encoding="utf-8")

            conformite_rc_result = verifier_conformite_rc(fichiers_pour_rc, rc_data)

            # Sauvegarder
            conformite_path = dossier_path / "conformite_rc.json"
            conformite_path.write_text(json.dumps(conformite_rc_result, ensure_ascii=False, indent=2), encoding="utf-8")
            fichiers_generes.append("conformite_rc.json")
            logger.info(
                f"Conformite RC: score={conformite_rc_result.get('score', '?')}/100, "
                f"conforme={conformite_rc_result.get('conforme', '?')}, "
                f"{len(conformite_rc_result.get('pieces_manquantes', []))} piece(s) manquante(s)"
            )
        else:
            logger.info("Conformite RC: pas de donnees RC disponibles, verification ignoree")
    except Exception as e:
        erreurs.append(f"Conformite RC: {e}")
        logger.error(f"Erreur conformite RC: {e}")

    # Feature : Coherence inter-documents (regex, pas d'appel API)
    coherence_result = None
    logger.info("Verification coherence inter-documents...")
    try:
        from auto_review import verifier_coherence_inter_documents

        # Charger tous les fichiers markdown generes
        fichiers_pour_coherence = {}
        for f_nom in fichiers_generes:
            if f_nom.endswith(".md"):
                f_path = dossier_path / f_nom
                if f_path.exists():
                    fichiers_pour_coherence[f_nom] = f_path.read_text(encoding="utf-8")

        if fichiers_pour_coherence:
            coherence_result = verifier_coherence_inter_documents(fichiers_pour_coherence)

            # Sauvegarder
            coherence_path = dossier_path / "coherence_check.json"
            coherence_path.write_text(json.dumps(coherence_result, ensure_ascii=False, indent=2), encoding="utf-8")
            fichiers_generes.append("coherence_check.json")
            logger.info(
                f"Coherence inter-documents: score={coherence_result.get('score', '?')}/100, "
                f"{len(coherence_result.get('incoherences', []))} incoherence(s)"
            )
    except Exception as e:
        erreurs.append(f"Coherence inter-documents: {e}")
        logger.error(f"Erreur coherence inter-documents: {e}")

    # Mettre a jour l'index
    _maj_index(dossier_nom, ao, dossier_path)

    success = len(fichiers_generes) >= 5  # Au moins 5 fichiers pour considerer le dossier complet
    result = {
        "success": success,
        "dossier_nom": dossier_nom,
        "fichiers": fichiers_generes,
        "nb_fichiers": len(fichiers_generes),
        "nb_mots": nb_mots_total,
        "erreurs": erreurs,
    }

    # Ajouter les infos de review, conformite RC et personnalisation au resultat
    if review_result:
        result["review"] = review_result
    if conformite_rc_result:
        result["conformite_rc"] = conformite_rc_result
    if coherence_result:
        result["coherence_inter_documents"] = coherence_result
    if personnalisation:
        result["personnalisation"] = {
            "type_acheteur": personnalisation["type_acheteur"],
        }

    if erreurs:
        result["avertissement"] = f"{len(erreurs)} erreur(s) lors de la generation"
        logger.warning(f"Dossier {dossier_nom}: {len(erreurs)} erreurs - {erreurs}")

    # Historique de generation
    try:
        import time as _time
        _log_generation(dossier_nom, ao, piece=None, nb_mots=nb_mots_total, duree_sec=0)
    except Exception:
        pass

    logger.info(f"Dossier complet genere: {dossier_nom} ({len(fichiers_generes)} fichiers, {nb_mots_total} mots)")
    return result


# Backward compat - l'ancienne fonction redirige vers la nouvelle
def generer_memoire_technique(ao: dict, type_presta: str = "Formation") -> dict:
    """Compatibilite arriere - genere maintenant un dossier complet."""
    return generer_dossier_complet(ao, type_presta)


# --- Regeneration piece par piece ---

# Mapping nom fichier -> (fonction_generatrice, args_supplementaires)
PIECES_GENERABLES = {
    "02_memoire_technique.md": "memoire",
    "03_lettre_candidature.md": "lettre",
    "04_bpu_dpgf.md": "bpu",
    "05_planning_previsionnel.md": "planning",
    "06_cv_formateurs.md": "cv",
    "07_dc1_dc2.md": "dc",
    "08_programme_formation.md": "programme",
    "09_references_clients.md": "references",
    "11_acte_engagement.md": "acte_engagement",
    "12_dume.md": "dume",
    "13_moyens_techniques.md": "moyens_techniques",
}


def regenerer_piece(dossier_nom: str, fichier: str, ao: dict = None, type_presta: str = "Formation") -> dict:
    """Regenere une seule piece d'un dossier existant.

    Args:
        dossier_nom: Nom du dossier (ex: AO_BOAMP-26_12345_20260404)
        fichier: Nom du fichier a regenerer (ex: 02_memoire_technique.md)
        ao: Dictionnaire de l'AO (si None, charge depuis fiche_ao.json)
        type_presta: Type de prestation

    Returns:
        dict avec success, fichier, nb_mots, erreur
    """
    import time as _time
    t0 = _time.time()

    piece_type = PIECES_GENERABLES.get(fichier)
    if not piece_type:
        return {"success": False, "erreur": f"Piece '{fichier}' non regenerable"}

    dossier_path = DOSSIERS_DIR / dossier_nom
    if not dossier_path.exists():
        return {"success": False, "erreur": f"Dossier '{dossier_nom}' introuvable"}

    # Charger l'AO depuis fiche_ao.json si pas fourni
    if not ao:
        fiche_path = dossier_path / "fiche_ao.json"
        if fiche_path.exists():
            ao = json.loads(fiche_path.read_text(encoding="utf-8"))
        else:
            return {"success": False, "erreur": "fiche_ao.json introuvable dans le dossier"}

    # Backup avant regeneration (pour diff visuel)
    ancien_path = dossier_path / fichier
    if ancien_path.exists():
        backup_path = ancien_path.with_suffix(".md.bak")
        backup_path.write_text(ancien_path.read_text(encoding="utf-8"), encoding="utf-8")

    try:
        generateurs = {
            "memoire": lambda: _generer_memoire(ao, type_presta),
            "lettre": lambda: _generer_lettre(ao),
            "bpu": lambda: _generer_bpu(ao),
            "planning": lambda: _generer_planning(ao),
            "cv": lambda: _generer_cv_formateurs(ao),
            "dc": lambda: _generer_dc1_dc2(ao),
            "programme": lambda: _generer_programme_formation(ao),
            "references": lambda: _generer_references_clients(ao),
            "acte_engagement": lambda: _generer_acte_engagement(ao),
            "dume": lambda: _generer_dume(ao),
            "moyens_techniques": lambda: _generer_moyens_techniques(ao),
        }

        contenu = generateurs[piece_type]()
        _sauvegarder(dossier_path, fichier, contenu)
        nb_mots = len(contenu.split())

        # Re-convertir en DOCX si applicable
        try:
            from export_docx_render import markdown_to_docx, DOCX_DISPONIBLE
            if DOCX_DISPONIBLE:
                docx_name = fichier.replace(".md", ".docx")
                titre_ao = ao.get("titre", "Appel d'offres")
                titre_doc = fichier.replace(".md", "").split("_", 1)[-1].replace("_", " ").title()
                markdown_to_docx(contenu, str(dossier_path / docx_name), titre_document=titre_ao, sous_titre=titre_doc)
                logger.info(f"DOCX re-genere: {docx_name}")
        except Exception as e:
            logger.warning(f"Erreur re-conversion DOCX: {e}")

        duree = round(_time.time() - t0, 1)

        # Logger dans l'historique
        _log_generation(dossier_nom, ao, fichier, nb_mots, duree)

        logger.info(f"Piece regeneree: {fichier} ({nb_mots} mots, {duree}s)")
        return {"success": True, "fichier": fichier, "nb_mots": nb_mots, "duree_sec": duree}

    except Exception as e:
        logger.error(f"Erreur regeneration {fichier}: {e}")
        return {"success": False, "erreur": str(e)}


# --- Historique des generations ---

GENERATIONS_LOG = DASHBOARD_DIR / "generations_log.json"


def _log_generation(dossier_nom: str, ao: dict, piece: str = None, nb_mots: int = 0, duree_sec: float = 0):
    """Enregistre une generation dans l'historique."""
    log = []
    if GENERATIONS_LOG.exists():
        try:
            log = json.loads(GENERATIONS_LOG.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            log = []

    # Estimation cout (Sonnet 4 : $3/MTok input, $15/MTok output)
    # ~1500 tokens input par appel, nb_mots * 1.5 tokens output
    tokens_input_est = 1500
    tokens_output_est = int(nb_mots * 1.5)
    cout_est = round((tokens_input_est * 3 + tokens_output_est * 15) / 1_000_000, 4)

    entry = {
        "timestamp": datetime.now().isoformat(),
        "dossier": dossier_nom,
        "ao_titre": ao.get("titre", "")[:100],
        "ao_id": ao.get("id", ""),
        "piece": piece or "dossier_complet",
        "nb_mots": nb_mots,
        "duree_sec": duree_sec,
        "cout_estime_usd": cout_est,
    }
    log.append(entry)

    # Garder les 500 dernieres entrees
    if len(log) > 500:
        log = log[-500:]

    GENERATIONS_LOG.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")


def charger_historique_generations() -> list:
    """Charge l'historique des generations."""
    if not GENERATIONS_LOG.exists():
        return []
    try:
        return json.loads(GENERATIONS_LOG.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []


def _convertir_en_docx(dossier_path: Path, ao: dict, fichiers_generes: list):
    """Convertit les documents Markdown cles en DOCX avec branding Almera."""
    try:
        from export_docx_render import markdown_to_docx, DOCX_DISPONIBLE
    except ImportError:
        logger.info("export_docx_render non disponible - pas de conversion DOCX")
        return

    if not DOCX_DISPONIBLE:
        logger.info("python-docx non installe - pas de conversion DOCX")
        return

    titre_ao = ao.get("titre", "Appel d'offres")

    # Mapping: nom fichier md -> titre du document DOCX
    documents_a_convertir = {
        "02_memoire_technique.md": "Memoire Technique",
        "03_lettre_candidature.md": "Lettre de Candidature",
        "04_bpu_dpgf.md": "Bordereau de Prix Unitaires - DPGF",
        "05_planning_previsionnel.md": "Planning Previsionnel",
        "06_cv_formateurs.md": "CV des Formateurs",
        "07_dc1_dc2.md": "DC1 - DC2",
        "08_programme_formation.md": "Programme de Formation",
        "09_references_clients.md": "References Clients",
        "11_acte_engagement.md": "Acte d'Engagement",
        "12_dume.md": "Document Unique de Marche Europeen",
        "13_moyens_techniques.md": "Moyens Techniques et Materiels",
    }

    nb_convertis = 0
    for md_filename, titre_doc in documents_a_convertir.items():
        if md_filename not in fichiers_generes:
            continue

        md_path = dossier_path / md_filename
        if not md_path.exists():
            continue

        docx_filename = md_filename.replace(".md", ".docx")
        docx_path = dossier_path / docx_filename

        try:
            contenu_md = md_path.read_text(encoding="utf-8")
            markdown_to_docx(
                contenu_md=contenu_md,
                output_path=docx_path,
                titre_document=titre_doc,
                sous_titre=titre_ao,
            )
            fichiers_generes.append(docx_filename)
            nb_convertis += 1
        except Exception as e:
            logger.warning(f"Erreur conversion DOCX {md_filename}: {e}")
            continue

    if nb_convertis:
        logger.info(f"{nb_convertis} document(s) converti(s) en DOCX")


def _sauvegarder(dossier_path: Path, nom_fichier: str, contenu: str):
    """Sauvegarde un fichier dans le dossier."""
    path = dossier_path / nom_fichier
    path.write_text(contenu, encoding="utf-8")


def _maj_index(dossier_nom: str, ao: dict, dossier_path: Path):
    """Met a jour dossiers_index.json."""
    index = []
    if DOSSIERS_INDEX.exists():
        try:
            index = json.loads(DOSSIERS_INDEX.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            index = []

    fichiers = [f.name for f in dossier_path.glob("*") if f.is_file()]
    entry = {
        "nom": dossier_nom,
        "ao_id": ao.get("id", ""),
        "ao_titre": ao.get("titre", ""),
        "nb_fichiers": len(fichiers),
        "fichiers": fichiers,
        "date_creation": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "pipeline_auto": True,
    }

    # Eviter les doublons
    index = [d for d in index if d.get("nom") != dossier_nom]
    index.insert(0, entry)

    DOSSIERS_INDEX.write_text(
        json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def valider_dossier_complet(ao: dict, dossier_path) -> dict:
    """Valide la completude et la coherence d'un dossier genere.

    Returns dict with:
    - completude_score: 0-100
    - pieces_presentes: list of files found
    - pieces_manquantes: list of expected but missing files
    - alertes: list of issues detected
    - pret_soumission: bool
    """
    dossier_path = Path(dossier_path)

    # Pieces attendues
    pieces_attendues = {
        'analyse_go_no_go.md': 'Analyse Go/No-Go',
        'memoire_technique.md': 'Memoire technique',
        'lettre_candidature.md': 'Lettre de candidature',
        'bpu_dpgf.md': 'BPU / DPGF',
        'bpu_dpgf.xlsx': 'DPGF Excel',
        'planning_previsionnel.md': 'Planning previsionnel',
        'cv_formateurs.md': 'CV Formateurs',
        'dc1_dc2.md': 'DC1 / DC2',
        'programme_formation.md': 'Programme de formation',
        'references_clients.md': 'References clients',
        'acte_engagement.md': 'Acte d\'engagement',
        'dume.md': 'DUME',
        'moyens_techniques.md': 'Moyens techniques',
        'checklist_soumission.md': 'Checklist soumission',
        'fiche_ao.json': 'Fiche AO',
    }

    presentes = []
    manquantes = []
    alertes = []

    for fichier, label in pieces_attendues.items():
        filepath = dossier_path / fichier
        if filepath.exists():
            presentes.append(fichier)
            # Verifier taille minimale
            taille = filepath.stat().st_size
            if taille < 100 and not fichier.endswith('.json'):
                alertes.append(f"{label} semble trop court ({taille} octets)")
        else:
            manquantes.append(fichier)

    # Score completude
    nb_total = len(pieces_attendues)
    nb_presentes = len(presentes)
    completude = round(nb_presentes / nb_total * 100) if nb_total > 0 else 0

    # Verifications croisees
    memoire_path = dossier_path / 'memoire_technique.md'
    if memoire_path.exists():
        memoire_txt = memoire_path.read_text(encoding='utf-8', errors='ignore')
        nb_mots = len(memoire_txt.split())
        if nb_mots < 2000:
            alertes.append(f"Memoire technique trop court ({nb_mots} mots, minimum recommande: 3000)")
        if 'conclusion' not in memoire_txt.lower():
            alertes.append("Memoire technique sans conclusion detectee")
        # Verifier que les formateurs du CV sont mentionnes
        cv_path = dossier_path / 'cv_formateurs.md'
        if cv_path.exists():
            cv_txt = cv_path.read_text(encoding='utf-8', errors='ignore')
            for f in ENTREPRISE['formateurs'][:3]:
                if f['nom'] in cv_txt and f['nom'] not in memoire_txt:
                    alertes.append(f"Formateur {f['nom']} dans CV mais pas mentionne dans le memoire")

    # Verifier coherence BPU vs budget
    bpu_path = dossier_path / 'bpu_dpgf.md'
    if bpu_path.exists():
        bpu_txt = bpu_path.read_text(encoding='utf-8', errors='ignore')
        budget = ao.get('budget_estime', 0) or 0
        if budget > 0:
            # Chercher le total dans le BPU
            totaux = re.findall(r'(\d[\d\s]*[\d])\s*(?:EUR|€)', bpu_txt)
            if totaux:
                try:
                    montant_max = max(int(t.replace(' ', '').replace('\u202f', '')) for t in totaux if int(t.replace(' ', '').replace('\u202f', '')) > 1000)
                    if montant_max > budget * 1.3:
                        alertes.append(f"BPU ({montant_max} EUR) depasse le budget de +{round((montant_max/budget - 1)*100)}%")
                except (ValueError, ZeroDivisionError):
                    pass

    # Verification deadline
    date_limite = ao.get('date_limite')
    if date_limite:
        try:
            dl = datetime.fromisoformat(str(date_limite).replace('Z', '+00:00'))
            jours = (dl.replace(tzinfo=None) - datetime.now()).days
            if jours < 0:
                alertes.append("ATTENTION: La deadline est depassee !")
            elif jours <= 2:
                alertes.append(f"URGENT: Deadline dans {jours} jour(s) !")
        except Exception:
            pass

    pret = completude >= 90 and not any('ATTENTION' in a or 'trop court' in a for a in alertes)

    return {
        'completude_score': completude,
        'pieces_presentes': presentes,
        'pieces_manquantes': manquantes,
        'alertes': alertes,
        'pret_soumission': pret,
        'nb_presentes': nb_presentes,
        'nb_total': nb_total,
    }
